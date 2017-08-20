from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, PatternFill
import os
import json
import copy

def add_cell(ws, cell_index, text, style):
	ws[cell_index] = text

	if style['alignment'] != None:
		ws[cell_index].alignment = style['alignment']
	if style['fill'] != None:
		ws[cell_index].fill=style['fill']
	if style['font'] != None:
		ws[cell_index].font=style['font']
	if style['border'] != None:
		ws[cell_index].border=style['border']


def add_row(ws, row_index, row, style, start_index=0):
	col_iterator = "A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z".split(',')
	index = start_index
	for x in row:
		cell_index = col_iterator[index]+str(row_index)
		index+=1
		add_cell(ws, cell_index, x, style)


def generate_xlsx_reports():
	wb = Workbook()	

	report_ws = wb.active
	report_ws.title = "Triage Summary"
	report_ws = wb.create_sheet("Allocation Recommendation")
	report_ws = wb.create_sheet("High Iterations")

	generate_triage_summary_report(wb)
	generate_allocation_report(wb)
	generate_high_iterations_report(wb)

	wb.save(os.path.join(os.path.dirname(__file__),'report/report.xlsx'))


def generate_high_iterations_report(wb):
	report = {}
	
	with open(os.path.join(os.path.dirname(__file__),'report/high_iterations_report.json'), 'rb') as fp:
		report = json.load(fp)
	
	report_ws = wb['High Iterations']
	
	#Format header
	report_ws.row_dimensions[1].height = 20

	#Format header
	report_ws.column_dimensions['A'].width = 16
	report_ws.column_dimensions['B'].width = 16
	report_ws.column_dimensions['C'].width = 16
	report_ws.column_dimensions['D'].width = 16
	report_ws.column_dimensions['E'].width = 16
	report_ws.column_dimensions['F'].width = 85
	report_ws.column_dimensions['G'].width = 30

	thin = Side(border_style="thin", color="000000")

	#Styles
	style = {
		"alignment": None,
		"fill": None,
		"font": None,
		"border": None
	}

	header_style = copy.deepcopy(style)
	header_style['alignment'] = Alignment(horizontal='center', vertical="center", wrap_text=True)
	header_style['fill'] = PatternFill("solid", fgColor="BBDEFB")
	header_style['font'] = Font(bold=True)
	header_style['border'] = Border(top=thin, bottom=thin, left=thin, right=thin)

	content_style = copy.deepcopy(style)
	header_style['alignment'] = Alignment(horizontal='center', vertical="center", wrap_text=True)

	header_row = [ 
					'Ticket_Number',
					'Customer',
					'Severity',
					'Category',
					'Assigned To',
					'Additional Info 1',
					'Additional Info 2'
				]
	add_row(report_ws, 1, header_row, style=header_style)

	row_index = 2
	for ticket in report:

		if len(report[ticket]['additional_info_1']) > 80:
					report_ws.row_dimensions[row_index].height = 40

		row = [
			ticket,
			report[ticket]['customer'],
			report[ticket]['severity'],
			report[ticket]['category'],
			report[ticket]['assigned_to']
		]

		add_row(report_ws, row_index, row, style=content_style)

		content_style['alignment'] = Alignment(horizontal='left', vertical="center", wrap_text=True)
		row = [
			report[ticket]['additional_info_1'],
			report[ticket]['additional_info_2']
		]

		add_row(report_ws, row_index, row, style=content_style, start_index=5)
		row_index+=1


def generate_allocation_report(wb):
	ticket_report = {}
	employee_report = {}
	with open(os.path.join(os.path.dirname(__file__),'report/employee_status_report.json'), 'rb') as fp:
		employee_report = json.load(fp)

	with open(os.path.join(os.path.dirname(__file__),'report/ticket_report.json'), 'rb') as fp:
		ticket_report = json.load(fp)

	report_ws = wb["Allocation Recommendation"]

	#Format header
	report_ws.row_dimensions[1].height = 20

	report_ws.column_dimensions['A'].width = 16
	report_ws.column_dimensions['B'].width = 16
	report_ws.column_dimensions['C'].width = 12
	report_ws.column_dimensions['D'].width = 10
	report_ws.column_dimensions['E'].width = 20
	report_ws.column_dimensions['F'].width = 17
	report_ws.column_dimensions['G'].width = 85
	report_ws.column_dimensions['H'].width = 17

	#Styles
	style = {
		"alignment": None,
		"fill": None,
		"font": None,
		"border": None
	}

	thin = Side(border_style="thin", color="000000")

	header_style = copy.deepcopy(style)
	header_style['alignment'] = Alignment(horizontal='center', vertical="center", wrap_text=True)
	header_style['fill'] = PatternFill("solid", fgColor="BBDEFB")
	header_style['font'] = Font(bold=True)
	header_style['border'] = Border(top=thin, bottom=thin, left=thin, right=thin)

	content_style = copy.deepcopy(style)
	header_style['alignment'] = Alignment(horizontal='center', vertical="center", wrap_text=True)

	header_row = [ 
					'Resource Name',
					'Ticket Number',
					'Customer',
					'Severity',
					'Category',
					'Status',
					'Triage Recommendation',
					'Last Worked By'
				]

	add_row(report_ws, 1, header_row, style=header_style)

	row_index = 2
	for employee in employee_report:
		if employee=="Unassigned":
			continue
		if len(employee_report[employee]['tickets']) > 0:
			ticket_list = employee_report[employee]['tickets'].split(';')
			n_tickets = len(ticket_list)
		
			report_ws.merge_cells('A'+str(row_index)+':A'+str(row_index+n_tickets-1))
			#Employee_name
			content_style['alignment'] = Alignment(horizontal='center', vertical="center", wrap_text=True)
			add_cell(report_ws, 'A'+str(row_index), employee, style=content_style)

			for ticket in ticket_list:
				t_info = ticket_report[ticket]

				style = copy.deepcopy(content_style)
				
				if t_info['backlog']:
					style['font'] = Font(color='FF0000')

				row = [
					ticket,
					t_info['customer'],
					t_info['severity'],
					t_info['category'],
					t_info['status']
				]

				style['alignment'] = Alignment(horizontal='center', vertical="center", wrap_text=True)
				add_row(report_ws, row_index, row, style=style, start_index=1)
				add_cell(report_ws, 'H'+str(row_index), t_info['last_worked_by'], style=style)

				style['alignment'] = Alignment(horizontal='left', vertical="center", wrap_text=True)

				if len(t_info['triage_recommendation']) > 100:
					report_ws.row_dimensions[row_index].height = 40

				add_cell(report_ws, 'G'+str(row_index), t_info['triage_recommendation'], style=style)
				row_index+=1

			# row_index+=1
			report_ws.merge_cells('A'+str(row_index)+':H'+str(row_index))

			style = copy.deepcopy(content_style)
			style['fill'] = PatternFill("solid", fgColor="FFFF00")
			style['alignment'] = Alignment(horizontal='center', vertical="center")
			double = Side(border_style="thin", color="000000")
			style['border'] = Border(top=double, left=double, right=double, bottom=double)
			add_cell(report_ws, 'A'+str(row_index), "Total "+str(n_tickets)+" tickets", style=style)
			row_index+=1

	if len(employee_report["Unassigned"]['tickets']) > 0:
		ticket_list = employee_report["Unassigned"]['tickets'].split(';')
		n_tickets = len(ticket_list)
	
		report_ws.merge_cells('A'+str(row_index)+':A'+str(row_index+n_tickets-1))
		#Employee_name
		content_style['alignment'] = Alignment(horizontal='center', vertical="center", wrap_text=True)
		add_cell(report_ws, 'A'+str(row_index), "Unassigned", style=content_style)

		for ticket in ticket_list:
			t_info = ticket_report[ticket]

			style = copy.deepcopy(content_style)
			
			if t_info['backlog']:
				style['font'] = Font(color='FF0000')

			row = [
				ticket,
				t_info['customer'],
				t_info['severity'],
				t_info['category'],
				t_info['status']
			]

			style['alignment'] = Alignment(horizontal='center', vertical="center", wrap_text=True)
			add_row(report_ws, row_index, row, style=style, start_index=1)
			add_cell(report_ws, 'H'+str(row_index), t_info['last_worked_by'], style=style)

			style['alignment'] = Alignment(horizontal='left', vertical="center", wrap_text=True)

			if len(t_info['triage_recommendation']) > 100:
				report_ws.row_dimensions[row_index].height = 40

			add_cell(report_ws, 'G'+str(row_index), t_info['triage_recommendation'], style=style)
			row_index+=1

		# row_index+=1
		report_ws.merge_cells('A'+str(row_index)+':H'+str(row_index))

		style = copy.deepcopy(content_style)
		style['fill'] = PatternFill("solid", fgColor="FFFF00")
		style['alignment'] = Alignment(horizontal='center', vertical="center")
		double = Side(border_style="thin", color="000000")
		style['border'] = Border(top=double, left=double, right=double, bottom=double)
		add_cell(report_ws, 'A'+str(row_index), "Total "+str(n_tickets)+" tickets", style=style)
		row_index+=1


def generate_triage_summary_report(wb):
	report = {}
	with open(os.path.join(os.path.dirname(__file__),'report/triager_summary_report.json'), 'rb') as fp:
		report = json.load(fp)

	report_ws = wb["Triage Summary"]

	#Format header
	#row merge
	report_ws.merge_cells('A1:A2')
	report_ws.merge_cells('F1:F2')
	report_ws.merge_cells('K1:K2')
	#column merge
	report_ws.merge_cells('B1:E1')
	report_ws.merge_cells('G1:J1')
	report_ws.merge_cells('L1:M1')
	report_ws.row_dimensions[1].height = 50
	# report_ws.row_dimensions[2].height = 10
	
	report_ws.column_dimensions['A'].width = 20
	report_ws.column_dimensions['B'].width = 10
	report_ws.column_dimensions['C'].width = 18
	report_ws.column_dimensions['D'].width = 8
	report_ws.column_dimensions['E'].width = 10
	report_ws.column_dimensions['F'].width = 14
	report_ws.column_dimensions['G'].width = 7
	report_ws.column_dimensions['H'].width = 7
	report_ws.column_dimensions['I'].width = 7
	report_ws.column_dimensions['J'].width = 7
	report_ws.column_dimensions['K'].width = 14
	report_ws.column_dimensions['L'].width = 8
	report_ws.column_dimensions['M'].width = 8

	#Styles

	style = {
		"alignment": None,
		"fill": None,
		"font": None,
		"border": None
	}

	thin = Side(border_style="thin", color="000000")

	style['alignment'] = Alignment(horizontal='center', vertical="center", wrap_text=True)
	style['fill'] = PatternFill("solid", fgColor="BBDEFB")
	style['font'] = Font(bold=True)
	style['border'] = Border(top=thin, bottom=thin, left=thin, right=thin)

	#row 1
	add_cell(report_ws, 'A1', "Triage Summary Date", style)
	add_cell(report_ws, 'B1', "# Cases Triaged/Allocated = "+str(report['total_allocated']), style)
	add_cell(report_ws, 'F1', "# of Priority Deliverables", style)
	add_cell(report_ws, 'G1', "Undelivered Work from Previous Days(Backlog)", style)
	add_cell(report_ws, 'K1', "# of Members available Today", style)
	add_cell(report_ws, 'L1', "# of New Maps", style)

	#row 2
	add_cell(report_ws, 'B2', "New Map", style)
	add_cell(report_ws, 'C2', "PER Map Change", style)
	add_cell(report_ws, 'D2', "Change", style)
	add_cell(report_ws, 'E2', "Research", style)
	add_cell(report_ws, 'G2', "Sev 1", style)
	add_cell(report_ws, 'H2', "Sev 2", style)
	add_cell(report_ws, 'I2', "Sev 3", style)
	add_cell(report_ws, 'J2', "Sev 4", style)
	add_cell(report_ws, 'L2', "B2B", style)
	add_cell(report_ws, 'M2', "B2Bi", style)

	style['fill'] = PatternFill("solid", fgColor="FFFFFF")
	style['font'] = Font(bold=False)

	row = [
		report['date'],
		report['category_report']['New Map'],
		report['category_report']['PER Map Change'],
		report['category_report']['Research'],
		report['category_report']['New Map'],
		report['priority_deliverables'],
		report['backlog_report']['Sev 1'],
		report['backlog_report']['Sev 2'],
		report['backlog_report']['Sev 3'],
		report['backlog_report']['Sev 4'],
		report['available_members'],
		report['new_map_report']['b2b'],
		report['new_map_report']['b2bi']		
	]

	add_row(report_ws, 3, row, style)